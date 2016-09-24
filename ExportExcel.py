from tkinter import messagebox
import ast
import openpyxl as xls
from openpyxl.styles import *
from openpyxl.chart import BarChart, BubbleChart, Reference, Series
from openpyxl.worksheet.datavalidation import DataValidation

import LoadInfos as li

idsColunas = ['B', 'C', 'D', 'E', 'F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO']

class ClassExcel:
    def __init__(self):
        self.LoadInfos = li.LoadInfos()
        self.portArea, self.portConcluidos, self.avalArea = self.LoadInfos.loadInfosExportExcel()
        self.excel = xls.Workbook()   #### Objeto excel sendo criado aqui
        self.ExcelCharts = ExcelAddTools(self.avalArea, self.areaPort, self.excel)
        
    def exportPort(self, listIds, areaPort):
        if len(portArea) == 0:
            messagebox.showwarning('Erro ao exportar!', 'Não é possível exportar um portfolio sem projetos em andamento')
        else: 
            gerenMaisProjs, tipo, idelinha = self.ExcelCharts.planilha(portArea)

            self.ExcelCharts.concluidos(portConcluidos, avalArea, areaPort, excel)

            self.ExcelCharts.grafGerenMaisProjs(gerenMaisProjs)

            self.ExcelCharts.grafConsolidado(listIds, portArea, avalArea, idelinha)

            gerMaiorCrit = maioresCrit(listIds, portArea, avalArea)

            for gerente in sorted(gerMaiorCrit.keys()):
                self.ExcelCharts.grafPorGerencia(gerente, gerMaiorCrit[gerente], portArea, avalArea, idelinha)
        
        #####  CRIAR POPUP DE CONFIRMAÇÃO
        excel.save('Portfolio ' + areaPort + ' - ' + tipo + '.xlsx')
        messagebox.showinfo('Portfolio exportado','O Portfolio ' + areaPort + ' foi exportado com sucesso!')

class ExcelAddTools:
    def __init__(self, avalArea, areaPort, excel):
        self.avalArea = avalArea
        self.areaPort = areaPort
        self.excel = excel
        self.LoadInfos = li.LoadInfos()
        
    #########  FUNÇÃO PARA GERAR PLANILHA  ############
    
    def planilha(self, portArea):
        ordemProjs = []
        
        projsXls = self.excel.get_sheet_by_name('Sheet')
        projsXls.title = 'Planilha.Projetos'
                    
        #########  PARA ORDENAR OS PROJETOS POR GERÊNCIA  #########
        for proj in portArea:                         
            if proj['gerencia'] not in ordemProjs: 
                ordemProjs.append(proj['gerencia'])
        #########  CRIAR PORTFOLIO  #########
                #### CABEÇALHO ####
                   # ESTILO #
        projsXls['B2'].fill = PatternFill(start_color='CCFFFF', fill_type = 'solid')
        for col in (idsColunas):
            projsXls[col + '3'].fill = PatternFill(start_color='99BBFF', fill_type = 'solid')
            projsXls[col + '3'].font = Font(name='Cambria', size = 11)
            projsXls[col + '3'].alignment = Alignment(horizontal='center')
            projsXls[col + '3'].border = Border(left=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                top=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'))
            projsXls[col + '4'].font = Font(name='Cambria', size = 11, bold = True)
            projsXls[col + '4'].alignment = Alignment(horizontal='center')
            projsXls[col + '4'].fill = PatternFill(start_color='DBDBDB', fill_type = 'solid')
            projsXls[col + '4'].border = Border(left=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                top=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='medium', color='FF000000'))
            projsXls[col + '2'].border = Border(top=Side(border_style='medium', color='FF000000'),
                                                bottom=Side(border_style='medium', color='FF000000'))
           # opções de borda {'dashed', 'hair', 'medium', 'mediumDashDotDot', 'dashDotDot', 'slantDashDot', 'thick', 'dashDot',    principais  -->'medium', 'thin'
           #, 'double', 'mediumDashDot', 'dotted', 'thin', 'mediumDashed'}
        projsXls['B3'].border = Border(left=Side(border_style='medium', color='FF000000'))
        projsXls['AO3'].border = Border(right=Side(border_style='medium', color='FF000000'))

                   # TÍTULOS #
        projsXls.merge_cells('B2:AO2')
        projsXls['B2'] = 'Portfolio'
        projsXls['B2'].font = Font(name='Cambria', size = 20)
        projsXls['B2'].alignment = Alignment(horizontal='center')
        projsXls['B2'].fill = PatternFill(start_color='E6E6FF', fill_type = 'solid')
        projsXls['B2'].border = Border(top=Side(border_style='medium', color='FF000000'),
                                       bottom=Side(border_style='medium', color='FF000000'),
                                       left=Side(border_style='medium', color='FF000000'))
        projsXls['AO2'].border = Border(top=Side(border_style='medium', color='FF000000'),
                                        bottom=Side(border_style='medium', color='FF000000'),
                                        right=Side(border_style='medium', color='FF000000'))


        projsXls['B3'] = '#'

        projsXls['C3'] = 'Solicitante'

        projsXls['D3'] = 'Gerência'

        projsXls['E3'] = 'Projetos'

        projsXls['F3'] = 'Área'

        projsXls['G3'] = 'Registro'

        projsXls['H3'] = 'Status'
        
        projsXls.merge_cells('I3:J3')
        projsXls['I3'] = 'Esforço do Projeto (HH)'
        
        projsXls.merge_cells('K3:L3')
        projsXls['K3'] = 'Confiabilidade nos fornecedores'
        
        projsXls.merge_cells('M3:N3')
        projsXls['M3'] = 'Habilidade interna dos profissionais'
        
        projsXls.merge_cells('O3:P3')
        projsXls['O3'] = 'Número de profissionais técnicos envolvidos no projeto'
        
        projsXls.merge_cells('Q3:R3')
        projsXls['Q3'] = 'Tendência de variação do escopo'

        projsXls.merge_cells('S3:T3')
        projsXls['S3'] = 'Comprometimento das partes interessadas'

        projsXls.merge_cells('U3:V3')
        projsXls['U3'] = 'Disponibilidade de recursos técnológicos'

        projsXls.merge_cells('W3:X3')
        projsXls['W3'] = 'Dificuldade técnica do projeto'
        
        projsXls.merge_cells('Y3:Z3')
        projsXls['Y3'] = 'Impacto no processo de negócio atual'

        projsXls['AA3'] = 'Nota Final'

        projsXls.merge_cells('AB3:AC3')
        projsXls['AB3'] = 'Relacionado a alguma atividade de negócio vital'
        
        projsXls.merge_cells('AD3:AE3')
        projsXls['AD3'] = 'Número de áreas negócio beneficiadas com o projeto'

        projsXls.merge_cells('AF3:AG3')
        projsXls['AF3'] = 'Relacionado as metas atuais da diretoria'

        projsXls.merge_cells('AH3:AI3')
        projsXls['AH3'] = 'ROI'

        projsXls.merge_cells('AJ3:AK3')
        projsXls['AJ3'] = 'Must Do'
        
        projsXls['AL3'] = 'Nota Final'
        
        projsXls['B4'].border = Border(left=Side(border_style='medium', color='FF000000'),
                                       right=Side(border_style='thin', color='FF000000'),
                                       top=Side(border_style='thin', color='FF000000'),
                                       bottom=Side(border_style='medium', color='FF000000'))
        projsXls['AO4'].border = Border(left=Side(border_style='thin', color='FF000000'),
                                       right=Side(border_style='medium', color='FF000000'),
                                       top=Side(border_style='thin', color='FF000000'),
                                       bottom=Side(border_style='medium', color='FF000000'))
                                       
        for nota in ('J','L','N','P','R','T','V','X','AA','AC','AE','AG','AI','AK'):
            projsXls[nota + '4'] = 'Nota'

        for peso in ('K','M','O','Q','S','U','W','Y','AB','AD','AF','AH','AJ','AL'):
            projsXls[peso + '4'] = 'Peso'

        projsXls['AA4'] = 'Complexidade'

        projsXls['AL4'] = 'Criticidade'

        projsXls['AM4'] = 'Esforço'

        projsXls['AN4'] = 'Budget'

        projsXls['AO4'] = 'Analista Funcional'

        ###########  CRIA DATA VALIDATION PARA AS NOTAS  #############
        listCompSetup = self.LoadInfos.loadXML(proj['tipo'], 'complexidade')
        listCritSetup = self.LoadInfos.loadXML(proj['tipo'], 'criticidade')
        
        ####  ESFORCO  ####
        setupEsf = self.LoadInfos.loadInfos('esforco', listCompSetup)
        validaEsf = DataValidation(type="list", formula1=str(setupEsf).replace('[','"').replace(']','"').replace("'",''), allow_blank=True)
        
        ####  CONFIABILIDADE  ####
        setupConf = self.LoadInfos.loadInfos('conf', listCompSetup)
        validaConf = DataValidation(type="list", formula1=str(setupConf).replace('[','"').replace(']','"').replace("'",''), allow_blank=True)

        ####  HABILIDADE  ####
        setupHab = self.LoadInfos.loadInfos('hab', listCompSetup)
        validaHab = DataValidation(type="list", formula1=str(setupHab).replace('[','"').replace(']','"').replace("'",''), allow_blank=True)

        #### NÚMERO DE PROFISSIONAIS  ####
        setupNumProfs = self.LoadInfos.loadInfos('numProfs', listCompSetup)
        validaNumProfs = DataValidation(type="list", formula1=str(setupNumProfs).replace('[','"').replace(']','"').replace("'",''), allow_blank=True)

        #### ESCOPO ####
        setupEscop = self.LoadInfos.loadInfos('escop', listCompSetup)
        validaEscop = DataValidation(type="list", formula1=str(setupEscop).replace('[','"').replace(']','"').replace("'",''), allow_blank=True)

        #### COMPROMETIMENTO ####
        setupComp = self.LoadInfos.loadInfos('comp', listCompSetup)
        validaComp = DataValidation(type="list", formula1=str(setupComp).replace('[','"').replace(']','"').replace("'",''), allow_blank=True)

        #### DISPONIBILIDADE TÉCNICA ####
        setupDispTec = self.LoadInfos.loadInfos('dispTec', listCompSetup)
        validaDispTec = DataValidation(type="list", formula1=str(setupDispTec).replace('[','"').replace(']','"').replace("'",''), allow_blank=True)

        #### DIFICULDADE TÉCNICA ####
        setupDifTec = self.LoadInfos.loadInfos('difTec', listCompSetup)
        validaDifTec = DataValidation(type="list", formula1=str(setupDifTec).replace('[','"').replace(']','"').replace("'",''), allow_blank=True)

        #### IMPACTO ####
        setupImpact = self.LoadInfos.loadInfos('impact', listCompSetup)
        validaImpact = DataValidation(type="list", formula1=str(setupImpact).replace('[','"').replace(']','"').replace("'",''), allow_blank=True)

        #### ATIVIDADE VITAL ####
        setupAtivVital = self.LoadInfos.loadInfos('ativVital', listCritSetup)
        validaAtivVital = DataValidation(type="list", formula1=str(setupAtivVital).replace('[','"').replace(']','"').replace("'",''), allow_blank=True)

        #### NÚMERO ÁREAS ####
        setupNumArea = self.LoadInfos.loadInfos('numArea', listCritSetup)
        validaNumArea = DataValidation(type="list", formula1=str(setupNumArea).replace('[','"').replace(']','"').replace("'",''), allow_blank=True)

        #### RELACIONADO AS METAS ####
        setupRelMetas = self.LoadInfos.loadInfos('relMetas', listCritSetup)
        validaRelMetas = DataValidation(type="list", formula1=str(setupRelMetas).replace('[','"').replace(']','"').replace("'",''), allow_blank=True)

        #### ROI ####
        setupRoi = self.LoadInfos.loadInfos('roi', listCritSetup)
        validaRoi = DataValidation(type="list", formula1=str(setupRoi).replace('[','"').replace(']','"').replace("'",''), allow_blank=True)

        #### MUST DO ####
        setupMustDo = self.LoadInfos.loadInfos('mustDo', listCritSetup)
        validaMustDo = DataValidation(type="list", formula1=str(setupMustDo).replace('[','"').replace(']','"').replace("'",''), allow_blank=True)
                
        #### PROJETOS ####
        linhaInfo = 5
        numProjGeren = []
        idelinha = {}
        for gerente in sorted(ordemProjs):
            numProj = 0
            for proj in portArea:
                if proj['gerencia'] == gerente:
                    numProj += 1
                    
                    projsXls['B' + str(linhaInfo)] = linhaInfo - 4
                    projsXls['C' + str(linhaInfo)] = proj['solicitante']
                    projsXls['D' + str(linhaInfo)] = proj['gerencia']
                    projsXls['E' + str(linhaInfo)] = proj['projeto']
                    projsXls['F' + str(linhaInfo)] = proj['subarea']
                    projsXls['G' + str(linhaInfo)] = proj['ticket']
                    projsXls['H' + str(linhaInfo)] = proj['status']

                    projsXls['AN' + str(linhaInfo)] = proj['budget']
                    projsXls['AO' + str(linhaInfo)] = proj['responsavel']

                    for col in (idsColunas):
                                projsXls[col + str(linhaInfo)].border = Border(left=Side(border_style='thin', color='FF000000'),
                                                                               right=Side(border_style='thin', color='FF000000'),
                                                                               top=Side(border_style='thin', color='FF000000'),
                                                                               bottom=Side(border_style='thin', color='FF000000'))

                    #######  ESCREVER AVALIAÇÃO  #######
                    for aval in self.avalArea:
                        if aval['id'] == proj['id']:
                            idelinha[aval['id']] = str(linhaInfo)
                            
                            projsXls.add_data_validation(validaEsf)
                            projsXls['I' + str(linhaInfo)] = aval["esforco"]
                            projsXls['J' + str(linhaInfo)] = int(aval["peso_esforco"])
                            
                            projsXls.add_data_validation(validaConf)
                            projsXls['K' + str(linhaInfo)] = aval["conf"]
                            projsXls['L' + str(linhaInfo)] = int(aval["peso_conf"])

                            projsXls.add_data_validation(validaHab)
                            projsXls['M' + str(linhaInfo)] = aval["hab"]
                            projsXls['N' + str(linhaInfo)] = int(aval["peso_hab"])

                            projsXls.add_data_validation(validaNumProfs)
                            projsXls['O' + str(linhaInfo)] = aval["numProfs"]
                            projsXls['P' + str(linhaInfo)] = int(aval["peso_numProfs"])

                            projsXls.add_data_validation(validaEscop)
                            projsXls['Q' + str(linhaInfo)] = aval["escop"]
                            projsXls['R' + str(linhaInfo)] = int(aval["peso_escop"])

                            projsXls.add_data_validation(validaComp)
                            projsXls['S' + str(linhaInfo)] = aval["comp"]
                            projsXls['T' + str(linhaInfo)] = int(aval["peso_comp"])

                            projsXls.add_data_validation(validaDispTec)
                            projsXls['U' + str(linhaInfo)] = aval["dispTec"]
                            projsXls['V' + str(linhaInfo)] = int(aval["peso_dispTec"])

                            projsXls.add_data_validation(validaDifTec)
                            projsXls['W' + str(linhaInfo)] = aval["difTec"]
                            projsXls['X' + str(linhaInfo)] = int(aval["peso_difTec"])

                            projsXls.add_data_validation(validaImpact)
                            projsXls['Y' + str(linhaInfo)] = aval["impact"]
                            projsXls['Z' + str(linhaInfo)] = int(aval["peso_impact"])

                            ###########  PARA ADICIONAR A LÓGICA DE CÁLCULO DO VALOR DE COMPLEXIDADE DENTRO DA CÉLULA  ###########

                            calcComp = '='
                            letra = 7
                            
                            for chave in ['esforco','conf','hab','numProfs','escop','comp','dispTec','difTec','impact']:
                                listaA = listCompSetup[chave]
                                calcComp += '('
                                letra += 1
                                for listaB in listaA:
                                    calcComp += 'IF(' + chr(letra + ord('A')) + str(linhaInfo) + ' = "' + listaB[0] + '",' + listaB[1] + ','
                                letra += 1
                                calcComp += '0))) * ' + chr(letra + ord('A')) + str(linhaInfo) + ') + '
                            calcComp += '0'

                            projsXls['AA' + str(linhaInfo)] = calcComp
                            ########################################################################################
                           
                            projsXls.add_data_validation(validaAtivVital)
                            projsXls['AB' + str(linhaInfo)] = aval["ativVital"]
                            projsXls['AC' + str(linhaInfo)] = int(aval["peso_ativVital"])
                            
                            projsXls.add_data_validation(validaNumArea)
                            projsXls['AD' + str(linhaInfo)] = aval["numArea"]
                            projsXls['AE' + str(linhaInfo)] = int(aval["peso_numArea"])

                            projsXls.add_data_validation(validaRelMetas)
                            projsXls['AF' + str(linhaInfo)] = aval["relMetas"]
                            projsXls['AG' + str(linhaInfo)] = int(aval["peso_relMetas"])

                            projsXls.add_data_validation(validaRoi)
                            projsXls['AH' + str(linhaInfo)] = aval["roi"]
                            projsXls['AI' + str(linhaInfo)] = int(aval["peso_roi"])

                            projsXls.add_data_validation(validaMustDo)
                            projsXls['AJ' + str(linhaInfo)] = aval["mustDo"]
                            projsXls['AK' + str(linhaInfo)] = int(aval["peso_mustDo"])

                            ###########  PARA ADICIONAR A LÓGICA DE CÁLCULO DO VALOR DE CRITICIDADE DENTRO DA CÉLULA  ###########
                            calcCrit = '='
                            letra = 0
                            
                            for chave in ['ativVital','numArea','relMetas','roi','mustDo']:
                                listaA = listCritSetup[chave]
                                calcCrit += '('
                                letra += 1
                                for listaB in listaA:
                                    calcCrit += 'IF(A' + chr(letra + ord('A')) + str(linhaInfo) + ' = "' + listaB[0] + '",' + listaB[1] + ','
                                letra += 1
                                calcCrit += '0))) * A' + chr(letra + ord('A')) + str(linhaInfo) + ') + '
                            calcCrit += '0'

                            projsXls['AL' + str(linhaInfo)] = calcCrit
                            ########################################################################################

                    linhaInfo += 1
                    
            numProjGeren.append(numProj)

        ###########  FIXA DATA VALIDATION PARA AS NOTAS  #############
        validaEsf.ranges.append('I5:I' + str(linhaInfo - 1))
        validaConf.ranges.append('K5:K' + str(linhaInfo - 1))
        validaHab.ranges.append('M5:M' + str(linhaInfo - 1))
        validaNumProfs.ranges.append('O5:O' + str(linhaInfo - 1))
        validaEscop.ranges.append('Q5:Q' + str(linhaInfo - 1))
        validaComp.ranges.append('S5:S' + str(linhaInfo - 1))
        validaDispTec.ranges.append('U5:U' + str(linhaInfo - 1))
        validaDifTec.ranges.append('W5:W' + str(linhaInfo - 1))
        validaImpact.ranges.append('Y5:Y' + str(linhaInfo - 1))
        validaAtivVital.ranges.append('AB5:AB' + str(linhaInfo - 1))
        validaNumArea.ranges.append('AD5:AD' + str(linhaInfo - 1))
        validaRelMetas.ranges.append('AF5:AF' + str(linhaInfo - 1))
        validaRoi.ranges.append('AH5:AH' + str(linhaInfo - 1))
        validaMustDo.ranges.append('AJ5:AJ' + str(linhaInfo - 1))

        ###########  DATA VALIDATION PARA OS PESOS  #############

        peso = DataValidation(type="list", formula1='"1,2,4,6"', allow_blank=True)
        projsXls.add_data_validation(peso)

        for letra in ('J','L','N','P','R','T','V','X','Z','AC','AE','AG','AI','AK'):
            peso.ranges.append(letra + '5' + ':' + letra + str(linhaInfo - 1))
            
        #########  CALCULA O NÚMERO DE PROJETOS RELACIONADOS A CADA GERENTE  ########        
        n = 0
        gerenMaisProjs = {}
        for gerente in sorted(ordemProjs):
            gerenMaisProjs[gerente] = str(numProjGeren[n])
            n += 1

        return gerenMaisProjs, proj['tipo'], idelinha

    #########  FUNÇÃO PARA GERAR PLANILHA DE PROJETOS CONCLUIDOS  ############
    
    def concluidos(portConcluidos, avalArea, areaPort, excel):
        ordemProjs = []

        projsXls = excel.create_sheet(title = 'Projetos em Produção')
                    
        #########  PARA ORDENAR OS PROJETOS POR GERÊNCIA  #########
        for proj in portConcluidos:                         
            if proj['gerencia'] not in ordemProjs: 
                ordemProjs.append(proj['gerencia'])
        #########  CRIAR PORTFOLIO  #########
                #### CABEÇALHO ####
                   # ESTILO #
        projsXls['B2'].fill = PatternFill(start_color='CCFFFF', fill_type = 'solid')
        for col in (idsColunas):
            projsXls[col + '3'].fill = PatternFill(start_color='99BBFF', fill_type = 'solid')
            projsXls[col + '3'].font = Font(name='Cambria', size = 11)
            projsXls[col + '3'].alignment = Alignment(horizontal='center')
            projsXls[col + '3'].border = Border(left=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                top=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='thin', color='FF000000'))
            projsXls[col + '4'].font = Font(name='Cambria', size = 11, bold = True)
            projsXls[col + '4'].alignment = Alignment(horizontal='center')
            projsXls[col + '4'].fill = PatternFill(start_color='DBDBDB', fill_type = 'solid')
            projsXls[col + '4'].border = Border(left=Side(border_style='thin', color='FF000000'),
                                                right=Side(border_style='thin', color='FF000000'),
                                                top=Side(border_style='thin', color='FF000000'),
                                                bottom=Side(border_style='medium', color='FF000000'))
            projsXls[col + '2'].border = Border(top=Side(border_style='medium', color='FF000000'),
                                                bottom=Side(border_style='medium', color='FF000000'))
           # opções de borda {'dashed', 'hair', 'medium', 'mediumDashDotDot', 'dashDotDot', 'slantDashDot', 'thick', 'dashDot',    principais  -->'medium', 'thin'
           #, 'double', 'mediumDashDot', 'dotted', 'thin', 'mediumDashed'}
        projsXls['B3'].border = Border(left=Side(border_style='medium', color='FF000000'))
        projsXls['AO3'].border = Border(right=Side(border_style='medium', color='FF000000'))

                   # TÍTULOS #
        projsXls.merge_cells('B2:AO2')
        projsXls['B2'] = 'Portfolio'
        projsXls['B2'].font = Font(name='Cambria', size = 20)
        projsXls['B2'].alignment = Alignment(horizontal='center')
        projsXls['B2'].fill = PatternFill(start_color='E6E6FF', fill_type = 'solid')
        projsXls['B2'].border = Border(top=Side(border_style='medium', color='FF000000'),
                                       bottom=Side(border_style='medium', color='FF000000'),
                                       left=Side(border_style='medium', color='FF000000'))
        projsXls['AO2'].border = Border(top=Side(border_style='medium', color='FF000000'),
                                        bottom=Side(border_style='medium', color='FF000000'),
                                        right=Side(border_style='medium', color='FF000000'))


        projsXls['B3'] = '#'

        projsXls['C3'] = 'Solicitante'

        projsXls['D3'] = 'Gerência'

        projsXls['E3'] = 'Projetos'

        projsXls['F3'] = 'Área'

        projsXls['G3'] = 'Registro'

        projsXls['H3'] = 'Status'
        
        projsXls.merge_cells('I3:J3')
        projsXls['I3'] = 'Esforço do Projeto (HH)'
        
        projsXls.merge_cells('K3:L3')
        projsXls['K3'] = 'Confiabilidade nos fornecedores'
        
        projsXls.merge_cells('M3:N3')
        projsXls['M3'] = 'Habilidade interna dos profissionais'
        
        projsXls.merge_cells('O3:P3')
        projsXls['O3'] = 'Número de profissionais técnicos envolvidos no projeto'
        
        projsXls.merge_cells('Q3:R3')
        projsXls['Q3'] = 'Tendência de variação do escopo'

        projsXls.merge_cells('S3:T3')
        projsXls['S3'] = 'Comprometimento das partes interessadas'

        projsXls.merge_cells('U3:V3')
        projsXls['U3'] = 'Disponibilidade de recursos técnológicos'

        projsXls.merge_cells('W3:X3')
        projsXls['W3'] = 'Dificuldade técnica do projeto'
        
        projsXls.merge_cells('Y3:Z3')
        projsXls['Y3'] = 'Impacto no processo de negócio atual'

        projsXls['AA3'] = 'Nota Final'

        projsXls.merge_cells('AB3:AC3')
        projsXls['AB3'] = 'Relacionado a alguma atividade de negócio vital'
        
        projsXls.merge_cells('AD3:AE3')
        projsXls['AD3'] = 'Número de áreas negócio beneficiadas com o projeto'

        projsXls.merge_cells('AF3:AG3')
        projsXls['AF3'] = 'Relacionado as metas atuais da diretoria'

        projsXls.merge_cells('AH3:AI3')
        projsXls['AH3'] = 'ROI'

        projsXls.merge_cells('AJ3:AK3')
        projsXls['AJ3'] = 'Must Do'
        
        projsXls['AL3'] = 'Nota Final'
        
        projsXls['B4'].border = Border(left=Side(border_style='medium', color='FF000000'),
                                       right=Side(border_style='thin', color='FF000000'),
                                       top=Side(border_style='thin', color='FF000000'),
                                       bottom=Side(border_style='medium', color='FF000000'))
        projsXls['AO4'].border = Border(left=Side(border_style='thin', color='FF000000'),
                                       right=Side(border_style='medium', color='FF000000'),
                                       top=Side(border_style='thin', color='FF000000'),
                                       bottom=Side(border_style='medium', color='FF000000'))
                                       
        for nota in ('J','L','N','P','R','T','V','X','AA','AC','AE','AG','AI','AK'):
            projsXls[nota + '4'] = 'Nota'

        for peso in ('K','M','O','Q','S','U','W','Y','AB','AD','AF','AH','AJ','AL'):
            projsXls[peso + '4'] = 'Peso'

        projsXls['AA4'] = 'Complexidade'

        projsXls['AL4'] = 'Criticidade'

        projsXls['AM4'] = 'Esforço'

        projsXls['AN4'] = 'Budget'

        projsXls['AO4'] = 'Analista Funcional'      
                
                #### PROJETOS ####
        linhaInfo = 5
        numProjGeren = []
        for gerente in sorted(ordemProjs):
            numProj = 0
            for proj in portConcluidos:
                if proj['gerencia'] == gerente:
                    numProj += 1
                    
                    projsXls['B' + str(linhaInfo)] = linhaInfo - 4
                    projsXls['C' + str(linhaInfo)] = proj['solicitante']
                    projsXls['D' + str(linhaInfo)] = proj['gerencia']
                    projsXls['E' + str(linhaInfo)] = proj['projeto']
                    projsXls['F' + str(linhaInfo)] = proj['subarea']
                    projsXls['G' + str(linhaInfo)] = proj['ticket']
                    projsXls['H' + str(linhaInfo)] = proj['status']

                    projsXls['AN' + str(linhaInfo)] = proj['budget']
                    projsXls['AO' + str(linhaInfo)] = proj['responsavel']

                    for col in (idsColunas):
                                projsXls[col + str(linhaInfo)].border = Border(left=Side(border_style='thin', color='FF000000'),
                                                                               right=Side(border_style='thin', color='FF000000'),
                                                                               top=Side(border_style='thin', color='FF000000'),
                                                                               bottom=Side(border_style='thin', color='FF000000'))

                    #######  ESCREVER AVALIAÇÃO  #######
                    for aval in avalArea:
                        if aval['id'] == proj['id']:
                            
                            projsXls['I' + str(linhaInfo)] = aval["esforco"]
                            projsXls['J' + str(linhaInfo)] = int(aval["peso_esforco"])
                            
                            projsXls['K' + str(linhaInfo)] = aval["conf"]
                            projsXls['L' + str(linhaInfo)] = int(aval["peso_conf"])

                            projsXls['M' + str(linhaInfo)] = aval["hab"]
                            projsXls['N' + str(linhaInfo)] = int(aval["peso_hab"])

                            projsXls['O' + str(linhaInfo)] = aval["numProfs"]
                            projsXls['P' + str(linhaInfo)] = int(aval["peso_numProfs"])

                            projsXls['Q' + str(linhaInfo)] = aval["escop"]
                            projsXls['R' + str(linhaInfo)] = int(aval["peso_escop"])

                            projsXls['S' + str(linhaInfo)] = aval["comp"]
                            projsXls['T' + str(linhaInfo)] = int(aval["peso_comp"])

                            projsXls['U' + str(linhaInfo)] = aval["dispTec"]
                            projsXls['V' + str(linhaInfo)] = int(aval["peso_dispTec"])

                            projsXls['W' + str(linhaInfo)] = aval["difTec"]
                            projsXls['X' + str(linhaInfo)] = int(aval["peso_difTec"])

                            projsXls['Y' + str(linhaInfo)] = aval["impact"]
                            projsXls['Z' + str(linhaInfo)] = int(aval["peso_impact"])

                            
                            projsXls['AA' + str(linhaInfo)] = aval["complexidade"]

                            projsXls['AB' + str(linhaInfo)] = aval["ativVital"]
                            projsXls['AC' + str(linhaInfo)] = int(aval["peso_ativVital"])
                            
                            projsXls['AD' + str(linhaInfo)] = aval["numArea"]
                            projsXls['AE' + str(linhaInfo)] = int(aval["peso_numArea"])

                            projsXls['AF' + str(linhaInfo)] = aval["relMetas"]
                            projsXls['AG' + str(linhaInfo)] = int(aval["peso_relMetas"])

                            projsXls['AH' + str(linhaInfo)] = aval["roi"]
                            projsXls['AI' + str(linhaInfo)] = int(aval["peso_roi"])

                            projsXls['AJ' + str(linhaInfo)] = aval["mustDo"]
                            projsXls['AK' + str(linhaInfo)] = int(aval["peso_mustDo"])
                            
                            projsXls['AL' + str(linhaInfo)] = aval['criticidade']

                    linhaInfo += 1
                    
            numProjGeren.append(numProj)

    #########  FUNÇÃO PARA GERAR GRÁFICO DE GERÊNTES COM MAIS PROJETOS SOLICITADOS  ############

    def grafGerenMaisProjs(gerenMaisProjs):
        global idsColunas
        
        grafGerens = excel.create_sheet(title = 'Projetos por Gerência')

        col = 3
        for gerente in sorted(gerenMaisProjs.keys()):
            grafGerens[idsColunas[col] + '10'] = gerente
            grafGerens[idsColunas[col] + '11'] = int(gerenMaisProjs[gerente])
            col += 1

        graf = BarChart()
        graf.type = "col"
        graf.style = 10
        graf.height = 13
        graf.width = 30
        graf.title = "Solicitações por gerência"

        gerente = Reference(grafGerens, min_col=4, min_row=10, max_col= (col + 1))
        numProjs = Reference(grafGerens, min_col=5, min_row=10, max_col= (col + 1), max_row=11)
        graf.add_data(numProjs, titles_from_data=True)
        graf.set_categories(gerente)
        graf.shape = 8
        grafGerens.add_chart(graf, "A1")

    #########  FUNÇÃO PARA GERAR GRÁFICO CONSOLIDADO DE PROJETOS  ############

    def grafConsolidado(listIds, portArea, avalArea, idelinha):
        
        grafConsol = excel.create_sheet(title = 'Consolidado')

        graf = BubbleChart()
        graf.style = 18
        graf.height = 13
        graf.width = 25
        graf.x_axis.scaling.min = 0
        graf.y_axis.scaling.min = 0
        graf.x_axis.scaling.max = 120
        graf.y_axis.scaling.max = 120
        graf.title = "Consolidado"

        grafConsol['M2'] = tipo
        grafConsol['N2'] = 'Complexidade'
        grafConsol['O2'] = 'Criticidade'
        grafConsol['P2'] = 'Custo'

        for celula in ('M2', 'N2', 'O2', 'P2'):
            grafConsol[celula].fill = PatternFill(start_color='99BBFF', fill_type = 'solid')
            grafConsol[celula].font = Font(name='Cambria', size = 11)
            grafConsol[celula].alignment = Alignment(horizontal='center')
            grafConsol[celula].border = Border(left=Side(border_style='medium', color='FF000000'),
                                               right=Side(border_style='medium', color='FF000000'),
                                               top=Side(border_style='medium', color='FF000000'),
                                               bottom=Side(border_style='medium', color='FF000000'))
        
        linha = 3
        for aval in avalArea:
            if aval['id'] in listIds:
                nomeProj = 'Não encontrado'
                for proj in portArea:
                    if aval['id'] == proj['id']:
                        nomeProj = proj['projeto']
                        
                        grafConsol['M' + str(linha)] = nomeProj
                        grafConsol['N' + str(linha)] = str('=Planilha.Projetos!AA' + idelinha[aval['id']])
                        grafConsol['O' + str(linha)] = str('=Planilha.Projetos!AL' + idelinha[aval['id']])
                        grafConsol['P' + str(linha)] = int(100)

                        for celula in ('M' + str(linha), 'N' + str(linha), 'O' + str(linha), 'P' + str(linha)):
                            grafConsol[celula].font = Font(name='Cambria', size = 11)
                            grafConsol[celula].border = Border(left=Side(border_style='thin', color='FF000000'),
                                                               right=Side(border_style='thin', color='FF000000'),
                                                               top=Side(border_style='thin', color='FF000000'),
                                                               bottom=Side(border_style='thin', color='FF000000'))
                        
                yvalues = Reference(grafConsol, min_col = 14, min_row = linha)
                xvalues = Reference(grafConsol, min_col = 15, min_row = linha)
                size = Reference(grafConsol, min_col = 16, min_row = linha)
                series = Series(values = yvalues, xvalues = xvalues, zvalues = size, title = nomeProj)

                graf.series.append(series)

                linha += 1

        grafConsol.add_chart(graf, 'A1')

    #########  FUNÇÃO PARA GERAR GRÁFICO DE MAIORES CRITICIDADES DOS PROJETOS  ############

    def maioresCrit(listIds, portArea, avalArea):
        global idsColunas

        grafMaiorCrit = excel.create_sheet(title = 'Maiores Pontuações')

        for col in ['A', 'B', 'C', 'D', 'E', 'F','G','H','I','J','K','L','M']:
            grafMaiorCrit[col + '24'].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                      bottom=Side(border_style='thin', color='FF000000'))
            grafMaiorCrit[col + '25'].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                      bottom=Side(border_style='thin', color='FF000000'))

        grafMaiorCrit['A24'] = 'Maiores Criticidades'
        grafMaiorCrit['A24'].font = Font(name='Calibri', size = 16, bold = True)
        grafMaiorCrit['A24'].alignment = Alignment(horizontal='center')
        grafMaiorCrit['A24'].fill = PatternFill(start_color='FFAD99', fill_type = 'solid')
        grafMaiorCrit['A24'].border = Border(left=Side(border_style='thin', color='FF000000'))
        grafMaiorCrit['A24'].border = Border(top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
        grafMaiorCrit['A25'].border = Border(top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
        grafMaiorCrit['N24'].border = Border(left=Side(border_style='thin', color='FF000000'))
        grafMaiorCrit.merge_cells('A24:M24')
        
        grafMaiorCrit['A25'] = 'Projeto'
        grafMaiorCrit['A25'].font = Font(name='Calibri', size = 11, italic = True)
        grafMaiorCrit['A25'].fill = PatternFill(start_color='DBDBDB', fill_type = 'solid')
        grafMaiorCrit['A25'].border = Border(left=Side(border_style='thin', color='FF000000'))
        grafMaiorCrit['J25'].border = Border(left=Side(border_style='thin', color='FF000000'),bottom=Side(border_style='thin', color='FF000000'))
        grafMaiorCrit.merge_cells('A25:H25')
        
        grafMaiorCrit['I25'] = 'Gerência'
        grafMaiorCrit['I25'].font = Font(name='Calibri', size = 11, italic = True)
        grafMaiorCrit['I25'].fill = PatternFill(start_color='DBDBDB', fill_type = 'solid')
        grafMaiorCrit['I25'].border = Border(left=Side(border_style='thin', color='FF000000'))
        grafMaiorCrit['L25'].border = Border(left=Side(border_style='thin', color='FF000000'))
        grafMaiorCrit.merge_cells('I25:K25')

        grafMaiorCrit['L25'] = 'Criticidade'
        grafMaiorCrit['L25'].font = Font(name='Calibri', size = 11, italic = True)
        grafMaiorCrit['L25'].fill = PatternFill(start_color='DBDBDB', fill_type = 'solid')
        grafMaiorCrit['N25'].border = Border(left=Side(border_style='thin', color='FF000000'))
        grafMaiorCrit.merge_cells('L25:M25')

        #########  SELECIONA A MAIOR CRITICIDADE  ###########
        gerMaiorCrit = {}
        for proj in portArea:     
            for aval in avalArea:
                if aval['id'] == proj['id']:
                    if proj['gerencia'] not in gerMaiorCrit.keys():
                        gerMaiorCrit[proj['gerencia']] = [aval['id'], proj['projeto'], aval['criticidade'], aval['complexidade']]
                    else:
                        maiorCrit = gerMaiorCrit[proj['gerencia']]
                        if int(maiorCrit[2]) < int(aval['criticidade']):
                            gerMaiorCrit[proj['gerencia']] = [aval['id'], proj['projeto'], aval['criticidade'], aval['complexidade']]
                        elif int(maiorCrit[2]) == int(aval['criticidade']):
                            if int(maiorCrit[3]) > int(aval['complexidade']):
                                gerMaiorCrit[proj['gerencia']] = [aval['id'], proj['projeto'], aval['criticidade'], aval['complexidade']]

        linha = 26
        for geren in sorted(gerMaiorCrit.keys()):
            infoProj = gerMaiorCrit[geren]
            
            #######  TABELA VISÍVEL ABAIXO DO GRÁFICO  ########
            grafMaiorCrit['I' + str(linha)] = geren
            grafMaiorCrit.merge_cells('I' + str(linha) + ':K' + str(linha))
            grafMaiorCrit['I' + str(linha)].font = Font(name='Cambria', size = 11)
            
            grafMaiorCrit['A' + str(linha)] = infoProj[1]
            grafMaiorCrit.merge_cells('A' + str(linha) + ':H' + str(linha))
            grafMaiorCrit['A' + str(linha)].font = Font(name='Cambria', size = 11)

            grafMaiorCrit['L' + str(linha)] = int(infoProj[2])
            grafMaiorCrit.merge_cells('L' + str(linha) + ':M' + str(linha))
            grafMaiorCrit['L' + str(linha)].font = Font(name='Cambria', size = 11)

            for col in ['A', 'B', 'C', 'D', 'E', 'F','G','H','I','J','K','L','M']:
                if col == 'A':
                    grafMaiorCrit[col + str(linha)].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                                    bottom=Side(border_style='thin', color='FF000000'),
                                                                    left=Side(border_style='thin', color='FF000000'))
                elif col == 'I':
                    grafMaiorCrit[col + str(linha)].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                                    bottom=Side(border_style='thin', color='FF000000'),
                                                                    left=Side(border_style='thin', color='FF000000'))
                elif col == 'L':
                    grafMaiorCrit[col + str(linha)].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                                    bottom=Side(border_style='thin', color='FF000000'),
                                                                    left=Side(border_style='thin', color='FF000000'))
                elif col == 'M':
                    grafMaiorCrit[col + str(linha)].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                                    bottom=Side(border_style='thin', color='FF000000'),
                                                                    right=Side(border_style='thin', color='FF000000'))
                else:
                    grafMaiorCrit[col + str(linha)].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                                    bottom=Side(border_style='thin', color='FF000000'))
            
            linha += 1
            

        graf = BarChart()
        graf.type = "col"
        graf.style = 10
        graf.title = "Maiores Criticidades"
        graf.height = 12
        graf.width = 30
        
        gerente = Reference(grafMaiorCrit, min_col=9, min_row=26, max_row=len(gerMaiorCrit.keys()) + 26)
        critProjs = Reference(grafMaiorCrit, min_col=12, min_row=25, max_row=len(gerMaiorCrit.keys()) + 25)
        graf.add_data(critProjs, titles_from_data=True)
        graf.set_categories(gerente)
        graf.shape = 4
        grafMaiorCrit.add_chart(graf, "A1")

        return gerMaiorCrit

    #########  FUNÇÃO PARA GERAR GRÁFICOS POR GERÊNCIA  ############

    def grafPorGerencia(gerente, gerMaiorCrit, portArea, avalArea, idelinha):
        global idsColunas
        
        grafPorGeren = excel.create_sheet(title = gerente)

        graf = BubbleChart()
        graf.style = 18
        graf.height = 11
        graf.width = 25
        graf.title = gerente
        graf.x_axis.scaling.min = 0
        graf.y_axis.scaling.min = 0
        graf.x_axis.scaling.max = 120
        graf.y_axis.scaling.max = 120

##        grafPorGeren.merge_cells('A1:N1')
##        grafPorGeren['A1'] = gerente
##        grafPorGeren['A1'].font = Font(name='Arial', size = 20)
##        grafPorGeren['A1'].alignment = Alignment(horizontal='center')
##        grafPorGeren['A1'].border = Border(top=Side(border_style='medium', color='FF000000'),
##                                                    bottom=Side(border_style='medium', color='FF000000'),
##                                                    left=Side(border_style='medium', color='FF000000'))
##        grafPorGeren['N1'].border = Border(top=Side(border_style='medium', color='FF000000'),
##                                                    bottom=Side(border_style='medium', color='FF000000'),
##                                                    right=Side(border_style='medium', color='FF000000'))
##        for letra in ['B', 'C', 'D', 'E', 'F','G','H','I','J','K','L','M']:
##            grafPorGeren[letra + '1'].border = Border(top=Side(border_style='medium', color='FF000000'),
##                                                      bottom=Side(border_style='medium', color='FF000000'))

        grafPorGeren['M2'] = tipo
        grafPorGeren['N2'] = 'Complexidade'
        grafPorGeren['O2'] = 'Criticidade'
        grafPorGeren['P2'] = 'Custo'

        for celula in ('M2', 'N2', 'O2', 'P2'):
            grafPorGeren[celula].fill = PatternFill(start_color='99BBFF', fill_type = 'solid')
            grafPorGeren[celula].font = Font(name='Cambria', size = 11)
            grafPorGeren[celula].alignment = Alignment(horizontal='center')
            grafPorGeren[celula].border = Border(left=Side(border_style='medium', color='FF000000'),
                                               right=Side(border_style='medium', color='FF000000'),
                                               top=Side(border_style='medium', color='FF000000'),
                                               bottom=Side(border_style='medium', color='FF000000'))
        
        grafPorGeren.merge_cells('A24:C24')
        grafPorGeren['A24'] = 'Chamado de maior criticidade:'
        grafPorGeren['A24'].font = Font(name='Calibri', size = 11, bold = True)
        grafPorGeren['A24'].fill = PatternFill(start_color='FFAD99', fill_type = 'solid')
        grafPorGeren['A24'].border = Border(top=Side(border_style='medium', color='FF000000'),
                                           bottom=Side(border_style='medium', color='FF000000'),
                                           left=Side(border_style='medium', color='FF000000'))
        grafPorGeren['B24'].border = Border(top=Side(border_style='medium', color='FF000000'),
                                                   bottom=Side(border_style='medium', color='FF000000'))
        grafPorGeren['C24'].border = Border(top=Side(border_style='medium', color='FF000000'),
                                                   bottom=Side(border_style='medium', color='FF000000'),
                                                   right=Side(border_style='medium', color='FF000000'))
                

        grafPorGeren.merge_cells('D24:I24')
        grafPorGeren['D24'] = gerMaiorCrit[1]
        grafPorGeren['D24'].font = Font(name='Calibri', size = 11)
        grafPorGeren['D24'].border = Border(top=Side(border_style='medium', color='FF000000'),
                                                     bottom=Side(border_style='medium', color='FF000000'),
                                                     left=Side(border_style='medium', color='FF000000'))
        grafPorGeren['I24'].border = Border(top=Side(border_style='medium', color='FF000000'),
                                                     bottom=Side(border_style='medium', color='FF000000'),
                                                     right=Side(border_style='medium', color='FF000000'))
        for letra in ['E', 'F','G','H']:
            grafPorGeren[letra + '24'].border = Border(top=Side(border_style='medium', color='FF000000'),
                                                       bottom=Side(border_style='medium', color='FF000000'))
        
        linha = 3
        for proj in portArea:
            if proj['gerencia'] == gerente:
                for aval in avalArea:
                    if aval['id'] == proj['id']:
                        grafPorGeren['M' + str(linha)] = proj['projeto']
                        grafPorGeren['N' + str(linha)] = str('=Planilha.Projetos!AA' + idelinha[aval['id']])
                        grafPorGeren['O' + str(linha)] = str('=Planilha.Projetos!AL' + idelinha[aval['id']])
                        grafPorGeren['P' + str(linha)] = int(100)

                        grafPorGeren['M' + str(linha)].font = Font(name='Cambria', size = 11)
                        grafPorGeren['M' + str(linha)].border = Border(left=Side(border_style='medium', color='FF000000'),
                                                             right=Side(border_style='medium', color='FF000000'),
                                                             top=Side(border_style='medium', color='FF000000'),
                                                             bottom=Side(border_style='medium', color='FF000000'))

                        for celula in ('N' + str(linha), 'O' + str(linha), 'P' + str(linha)):
                            grafPorGeren[celula].font = Font(name='Cambria', size = 11)
                            grafPorGeren[celula].alignment = Alignment(horizontal='center')
                            grafPorGeren[celula].border = Border(right=Side(border_style='medium', color='FF000000'),
                                                                 top=Side(border_style='medium', color='FF000000'),
                                                                 bottom=Side(border_style='medium', color='FF000000'))
                        
                        yvalues = Reference(grafPorGeren, min_col = 14, min_row = linha)
                        xvalues = Reference(grafPorGeren, min_col = 15, min_row = linha)
                        size = Reference(grafPorGeren, min_col = 16, min_row = linha)
                        series = Series(values = yvalues, xvalues = xvalues, zvalues = size, title = proj['projeto'])

                        graf.series.append(series)

                        linha += 1

        grafPorGeren.add_chart(graf, 'A2')
